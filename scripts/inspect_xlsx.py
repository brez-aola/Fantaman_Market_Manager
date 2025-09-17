import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as e:
    print("MISSING_OPENPYXL")
    sys.exit(2)

p = (
    Path(sys.argv[1])
    if len(sys.argv) > 1
    else Path(r"C:/work/fantacalcio/Rose_fantalega-darko-pancev.xlsx")
)
if not p.exists():
    print("FILE_NOT_FOUND", p)
    sys.exit(3)

wb = load_workbook(p, read_only=True)
print("SHEETS:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print("\n--- SHEET:", name, "---")
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        print("  (empty)")
        continue
    print("HEADER:", header)
    print("FIRST 10 ROWS:")
    for i, r in enumerate(rows):
        print(i + 1, r)
        if i >= 9:
            break

print("\nDone")
