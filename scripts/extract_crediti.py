from pathlib import Path
from openpyxl import load_workbook
import re
import json

XLSX = Path("Rose_fantalega-darko-pancev.xlsx")
if not XLSX.exists():
    print("xlsx not found:", XLSX)
    raise SystemExit(1)

wb = load_workbook(XLSX, read_only=True)
if "TutteLeRose" not in wb.sheetnames:
    print("sheet missing")
    raise SystemExit(2)
ws = wb["TutteLeRose"]
rows = list(ws.iter_rows(values_only=True))
# team names row is index 4 (1-based row 5)
team_row = rows[4]
team_starts = [
    (idx, cell.strip())
    for idx, cell in enumerate(team_row)
    if isinstance(cell, str) and cell.strip()
]
print("Detected teams (col, name):")
for t in team_starts:
    print(" ", t)

crediti_re = re.compile(r"crediti\s*residui\s*[:\-]?\s*(\d+)", re.I)
proposed = {}
for col, name in team_starts:
    last = None
    for r in rows:
        v = None
        try:
            v = r[col]
        except Exception:
            v = None
        if isinstance(v, str):
            m = crediti_re.search(v)
            if m:
                last = int(m.group(1))
    if last is not None:
        proposed[name] = last
    else:
        proposed[name] = None

print("\nProposed credits mapping:")
for k, v in proposed.items():
    print(f"  {k}:", v)

# write mapping to a small JSON file for later import if you confirm
Path("scripts/crediti_proposed.json").write_text(json.dumps(proposed, indent=2))
print("\nWrote scripts/crediti_proposed.json")
