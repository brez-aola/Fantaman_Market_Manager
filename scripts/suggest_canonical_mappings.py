"""Generate suggested canonical team mappings from roster XLSX files.

Improvements over the original script:
 - Identify columns whose header contains "squadra", "team" or "rosa" and extract only that column
 - Filter out obvious non-team strings (URLs, headers, numeric-only, known labels)
 - Produce three CSVs:
   * suggested_canonical_mappings_all.csv
   * suggested_canonical_mappings_highconf.csv (score >= threshold)
   * suggested_canonical_mappings_lowconf.csv  (score < threshold)
 - Parameterized fuzzy threshold (default 0.60)

Uses RapidFuzz if installed, otherwise difflib.get_close_matches.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List, Tuple

try:
    from rapidfuzz import process as rf_process
except Exception:
    rf_process = None

from difflib import get_close_matches
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "giocatori.db"
OUT_ALL = ROOT / "suggested_canonical_mappings_all.csv"
OUT_HIGH = ROOT / "suggested_canonical_mappings_highconf.csv"
OUT_LOW = ROOT / "suggested_canonical_mappings_lowconf.csv"

EXCEL_FILES = [
    ROOT / "Rose_fantalega-darko-pancev.xlsx",
    ROOT / "lista_calciatori_svincolati_classic_fantalega-darko-pancev.xlsx",
]

# optional list of real clubs to exclude from suggestions
REAL_CLUBS_FILE = ROOT / "data" / "real_clubs.txt"


BLACKLIST_WORDS = {
    "ruolo",
    "crediti",
    "rose",
    "rose lega",
    "calciatori",
    "rose lega fantalega",
    "rose lega fantalega darko",
}


def is_probable_team_name(s: str) -> bool:
    if not s:
        return False
    s_stripped = s.strip()
    if len(s_stripped) < 2:
        return False
    low = s_stripped.lower()
    if any(b in low for b in BLACKLIST_WORDS):
        return False
    if low.startswith("http") or ".com" in low or ".it" in low:
        return False
    # numeric-like lines
    if all(ch.isdigit() or ch in " .," for ch in low):
        return False
    # exclude common header labels
    if low in {"nome", "ruolo", "squadra", "crediti residui", "par"}:
        return False
    # exclude very short aliases (likely abbreviations or headers)
    if len(low) < 4:
        return False
    return True


def extract_team_names_from_roster(path: Path) -> List[str]:
    wb = load_workbook(path, read_only=True)
    names = set()
    for ws in wb.worksheets:
        # find header row and candidate columns
        header_row = None
        for r in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            if any(isinstance(c, str) and ("squadra" in c.lower() or "team" in c.lower() or "rosa" in c.lower()) for c in r if c):
                header_row = r
                break
        if header_row:
            # find indices with likely team header
            idxs = [i for i, c in enumerate(header_row) if c and isinstance(c, str) and ("squadra" in c.lower() or "team" in c.lower() or "rosa" in c.lower())]
            if not idxs:
                continue
            # read values from that column
            for row in ws.iter_rows(min_row=2, values_only=True):
                for i in idxs:
                    try:
                        v = row[i]
                    except Exception:
                        v = None
                    if v and isinstance(v, str) and is_probable_team_name(v):
                        names.add(v.strip())
            # we processed a sheet with team column; stop after first match
            break
        else:
            # fallback: scan first 200 rows for probable team-like short strings
            for row in ws.iter_rows(min_row=1, max_row=200, max_col=6, values_only=True):
                for v in row:
                    if v and isinstance(v, str) and len(v) < 80 and is_probable_team_name(v):
                        names.add(v.strip())
            # fallback only first sheet
            break
    return sorted(names)


def load_existing_teams(session) -> List[str]:
    from app.models import Team, CanonicalMapping

    teams = [t.name for t in session.query(Team).all()]
    # include canonical mapping targets too
    try:
        for cm in session.query(CanonicalMapping).all():
            if cm.canonical not in teams:
                teams.append(cm.canonical)
    except Exception:
        pass
    return teams


def best_match(name: str, choices: List[str]) -> Tuple[str, float]:
    if not choices:
        return ("", 0.0)
    if rf_process:
        res = rf_process.extractOne(name, choices)
        if res:
            return (res[0], res[1] / 100.0)
    else:
        res = get_close_matches(name, choices, n=1, cutoff=0.0)
        if res:
            return (res[0], 1.0)
    return ("", 0.0)


def main(threshold: float = 0.6):
    engine = create_engine(f"sqlite:///{DB_PATH}")
    Session = sessionmaker(bind=engine)
    session = Session()

    all_aliases = set()
    for f in EXCEL_FILES:
        if f.exists():
            names = extract_team_names_from_roster(f)
            all_aliases.update(names)
        else:
            print(f"Warning: roster file not found: {f}")

    choices = load_existing_teams(session)

    # load real clubs if present
    real_clubs = set()
    if REAL_CLUBS_FILE.exists():
        for ln in REAL_CLUBS_FILE.read_text(encoding='utf-8').splitlines():
            ln = ln.strip()
            if ln:
                real_clubs.add(ln)

    rows = []
    for a in sorted(all_aliases):
        match_name, score = best_match(a, choices)
        note = "rapidfuzz" if rf_process else "difflib"
        # filter out suggestions whose best_match is a real club
        if match_name in real_clubs:
            continue
        rows.append((a, match_name, float(score), note))

    # write all
    with OUT_ALL.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["source_alias", "best_match", "score", "note"])
        for r in rows:
            writer.writerow([r[0], r[1], f"{r[2]:.3f}", r[3]])

    # split by threshold
    with OUT_HIGH.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["source_alias", "best_match", "score", "note"])
        for r in rows:
            if r[2] >= threshold:
                writer.writerow([r[0], r[1], f"{r[2]:.3f}", r[3]])

    with OUT_LOW.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["source_alias", "best_match", "score", "note"])
        for r in rows:
            if r[2] < threshold:
                writer.writerow([r[0], r[1], f"{r[2]:.3f}", r[3]])

    print(f"Wrote {OUT_ALL}, {OUT_HIGH}, {OUT_LOW} (threshold={threshold})")


if __name__ == '__main__':
    main()
