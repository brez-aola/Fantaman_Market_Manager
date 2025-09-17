import sys
from pathlib import Path
import sqlite3
from app.db import get_connection
from openpyxl import load_workbook

DB = Path("/mnt/c/work/fantacalcio/giocatori.db")
XLSX = (
    Path(sys.argv[1])
    if len(sys.argv) > 1
    else Path("/mnt/c/work/fantacalcio/Rose_fantalega-darko-pancev.xlsx")
)

if not XLSX.exists():
    print("xlsx not found:", XLSX)
    sys.exit(2)
if not DB.exists():
    print("db not found:", DB)
    sys.exit(3)

wb = load_workbook(str(XLSX), read_only=True)
if "TutteLeRose" not in wb.sheetnames:
    print("sheet TutteLeRose not present; sheets:", wb.sheetnames)
    sys.exit(4)
ws = wb["TutteLeRose"]

# read rows as list for easier indexing
rows = list(ws.iter_rows(values_only=True))
if len(rows) < 6:
    print("unexpected sheet format (too few rows)")
    sys.exit(5)

# team names are in row index 4 (1-based row 5) per this workbook
team_row = rows[4]
team_starts = []  # list of (start_col, team_name)
for idx, cell in enumerate(team_row):
    if cell and isinstance(cell, str) and cell.strip():
        name = cell.strip()
        team_starts.append((idx, name))

if not team_starts:
    print("no team names found in expected row 4")
    sys.exit(6)

print("Found teams (start_col, name):")
for s in team_starts:
    print(" ", s)

# For each team, read columns start..start+3 and rows from row index 5 onward (0-based index 5 = sheet row 6)
team_players = {name: [] for (_, name) in team_starts}
start_cols = [s for (s, _) in team_starts]
start_cols_sorted = sorted(start_cols)

# We'll iterate rows starting at index 5 (0-based)
for r in rows[5:]:
    for start, tname in team_starts:
        # try to read 4 columns: role, calciatore, squadra, costo
        try:
            role = r[start]
            calciatore = r[start + 1]
            squadra_reale = r[start + 2]
            costo = r[start + 3]
        except IndexError:
            role = calciatore = squadra_reale = costo = None
        if calciatore and isinstance(calciatore, str) and calciatore.strip():
            name = calciatore.strip()
            # normalize cost to float or 0
            try:
                c = float(str(costo).strip()) if costo not in (None, "") else 0.0
            except Exception:
                try:
                    # sometimes cost may be like '1 ' or with comma
                    c = float(str(costo).replace(",", ".").strip())
                except Exception:
                    c = 0.0
            team_players[tname].append(
                {
                    "Nome": name,
                    "Ruolo": (role or "").strip() if role else "",
                    "Sq.": (squadra_reale or "").strip() if squadra_reale else "",
                    "Costo": c,
                }
            )

# Summary counts
total_players = sum(len(v) for v in team_players.values())
print("\nParsed players total:", total_players)
for t, lst in team_players.items():
    print(" ", t, len(lst))

# Now open DB and perform inserts/updates
conn = get_connection(str(DB))
# row_factory already set by helper
cur = conn.cursor()
# fetch columns of giocatori
cur.execute("PRAGMA table_info(giocatori)")
cols_info = cur.fetchall()
cols = [c[1] for c in cols_info]
print("\nDB giocatori columns:", cols)

inserted = 0
updated = 0
notfound = 0

for team, players in team_players.items():
    for p in players:
        nome = p["Nome"]
        ruolo = p["Ruolo"]
        sqreal = p["Sq."]
        costo = p["Costo"]
        # Try to find existing by Nome (case-insensitive)
        cur.execute(
            'SELECT rowid, * FROM giocatori WHERE lower("Nome") = lower(?)', (nome,)
        )
        row = cur.fetchone()
        if row:
            rowid = (
                row["rowid"]
                if "rowid" in row.keys()
                else row["id"] if "id" in row.keys() else None
            )
            # Update fields: squadra, Costo, anni_contratto=1, opzione='NO', "Sq." and "R." and "Nome" possibly
            updates = {
                "squadra": team,
                "Costo": costo,
                "anni_contratto": 1,
                "opzione": "NO",
                "Sq.": sqreal,
                "R.": ruolo,
            }
            set_clause = ", ".join([f'"{k}" = ?' for k in updates.keys()])
            params = list(updates.values()) + [rowid]
            try:
                cur.execute(f"UPDATE giocatori SET {set_clause} WHERE rowid=?", params)
                updated += 1
            except Exception as e:
                print("Failed to update", nome, e)
        else:
            # Insert new row: attempt to set common columns present, fallback to minimal
            to_insert = {}
            for k, v in [
                ("Nome", nome),
                ("Sq.", sqreal),
                ("R.", ruolo),
                ("Costo", costo),
                ("squadra", team),
                ("anni_contratto", 1),
                ("opzione", "NO"),
            ]:
                if k in cols:
                    to_insert[k] = v
            if not to_insert:
                print("No suitable columns to insert for", nome)
                notfound += 1
                continue
            col_names = ", ".join([f'"{c}"' for c in to_insert.keys()])
            qmarks = ", ".join(["?"] * len(to_insert))
            try:
                cur.execute(
                    f"INSERT INTO giocatori ({col_names}) VALUES ({qmarks})",
                    list(to_insert.values()),
                )
                inserted += 1
            except Exception as e:
                print("Failed to insert", nome, e)

# After changes, recompute team cash: for each team set cassa_attuale = cassa_iniziale - SUM(Costo of assigned players)
print("\nUpdating team cash balances...")
for team in team_players.keys():
    # ensure fantateam row exists
    cur.execute(
        "SELECT cassa_iniziale, cassa_attuale FROM fantateam WHERE squadra=?", (team,)
    )
    r = cur.fetchone()
    if r and r["cassa_iniziale"] is not None:
        starting = float(r["cassa_iniziale"])
    else:
        starting = 300.0
    cur.execute(
        'SELECT COALESCE(SUM(CAST(REPLACE(REPLACE(REPLACE(COALESCE("Costo", "0"), ",", ""), "%", ""), " ", "") AS REAL)),0) as spent FROM giocatori WHERE squadra=? AND NOT (opzione = "SI" AND anni_contratto IS NULL)',
        (team,),
    )
    spent_row = cur.fetchone()
    spent = (
        float(spent_row["spent"])
        if spent_row and spent_row["spent"] is not None
        else 0.0
    )
    new_attuale = starting - spent
    cur.execute(
        "INSERT OR REPLACE INTO fantateam(squadra, carryover, cassa_iniziale, cassa_attuale) VALUES (?,?,?,?)",
        (team, 0, starting, new_attuale),
    )
    print(
        " Team",
        team,
        "starting=",
        starting,
        "spent=",
        spent,
        "new cassa_attuale=",
        new_attuale,
    )

conn.commit()
conn.close()

print("\nSummary: inserted=", inserted, "updated=", updated, "skipped=", notfound)
