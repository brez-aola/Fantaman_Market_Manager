from pathlib import Path

from openpyxl import load_workbook


def parse_roster(path_or_file):
    """Parse the Rose_fantalega-*.xlsx workbook and return a mapping {team_name: [player dicts]}.

    This mirrors the logic in scripts/import_roster.py but is reusable for the admin UI.
    """
    p = Path(path_or_file)
    wb = load_workbook(str(p), read_only=True)
    if "TutteLeRose" not in wb.sheetnames:
        raise ValueError("sheet TutteLeRose not present")
    ws = wb["TutteLeRose"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError("unexpected sheet format (too few rows)")

    team_row = rows[4]
    team_starts = []
    for idx, cell in enumerate(team_row):
        if cell and isinstance(cell, str) and cell.strip():
            name = cell.strip()
            team_starts.append((idx, name))
    if not team_starts:
        raise ValueError("no team names found in expected row 4")

    team_players = {name: [] for (_, name) in team_starts}

    for r in rows[5:]:
        for start, tname in team_starts:
            try:
                role = r[start]
                calciatore = r[start + 1]
                squadra_reale = r[start + 2]
                costo = r[start + 3]
            except IndexError:
                role = calciatore = squadra_reale = costo = None
            if calciatore and isinstance(calciatore, str) and calciatore.strip():
                name = calciatore.strip()
                try:
                    c = float(str(costo).strip()) if costo not in (None, "") else 0.0
                except (ValueError, TypeError):
                    try:
                        c = float(str(costo).replace(",", ".").strip())
                    except (ValueError, TypeError):
                        c = 0.0
                team_players[tname].append(
                    {
                        "Nome": name,
                        "Ruolo": (
                            ((role or "").strip()[:1].upper().replace("G", "P"))
                            if role
                            else ""
                        ),
                        "Sq.": (squadra_reale or "").strip() if squadra_reale else "",
                        "Costo": c,
                    }
                )

    # simple validations: ensure each team has at least one player and costs are non-negative
    issues = []
    for t, lst in team_players.items():
        if not lst:
            issues.append((t, "no players found"))
        for p in lst:
            if p.get("Costo", 0) < 0:
                issues.append((t, f"negative cost for {p.get('Nome')}"))
    return {"teams": team_players, "issues": issues}


def apply_roster(db_path, team_players, audit_info: dict | None = None):
    """Apply parsed roster data to the sqlite DB using the same logic as the script.

    Returns summary dict with inserted/updated counts.
    """
    # Use SQLAlchemy ORM for data updates and audit
    from sqlalchemy import create_engine, func, text
    from sqlalchemy.orm import sessionmaker

    from app.models import ImportAudit
    import logging
    from sqlalchemy.exc import SQLAlchemyError

    engine = create_engine(f"sqlite:///{db_path}")
    Session = sessionmaker(bind=engine)
    s = Session()

    inserted = 0
    updated = 0
    notfound = 0

    try:
        # ensure Player model maps to giocatori table; fallback to raw SQL if Player has different schema
        for team_name, players in team_players.items():
            # resolve or create Team if necessary

            # find Team by exact name; create if missing
            team = s.query(func.count()).select_from(text("teams")) if False else None
            # simple ORM attempt
            try:
                from app.models import Team

                team = s.query(Team).filter(Team.name == team_name).first()
                if not team:
                    # create lightweight team entry (cash default 0)
                    team = Team(name=team_name, cash=0)
                    s.add(team)
                    s.flush()
            except Exception as e:
                # Be explicit about failures resolving/creating Team; continue with raw SQL path
                logging.exception("Failed to resolve or create Team '%s': %s", team_name, e)
                team = None

            for p in players:
                nome = p["Nome"]
                ruolo = p["Ruolo"]
                sqreal = p["Sq."]
                costo = p["Costo"]

                # try to find existing player row in giocatori by nome (case-insensitive)
                # We will operate via raw SQL for giocatori table for full compatibility
                existing = s.execute(
                    text(
                        'SELECT rowid FROM giocatori WHERE lower("Nome") = lower(:nome)'
                    ),
                    {"nome": nome},
                ).fetchone()
                if existing:
                    rowid = existing[0]
                    # Update via SQL to preserve schema differences
                    updates = {
                        "squadra": team_name,
                        "Costo": costo,
                        "anni_contratto": 1,
                        "opzione": "NO",
                        "Sq.": sqreal,
                        "R.": (ruolo[:1].upper().replace("G", "P") if ruolo else ""),
                    }
                    set_clause = ", ".join([f'"{k}" = :{k}' for k in updates.keys()])
                    # Validate that update columns exist in the giocatori table
                    cur_cols = [r[1] for r in s.execute(text("PRAGMA table_info(giocatori)")).fetchall()]
                    for k in updates.keys():
                        if k not in cur_cols:
                            raise ValueError(f"Unexpected column name for giocatori update: {k}")

                    params = {**updates, "rowid": rowid}
                    try:
                        # values are parameterized; identifiers were validated above
                        s.execute(text("UPDATE giocatori SET " + set_clause + " WHERE rowid=:rowid"), params)  # nosec: B608 - identifiers validated, values parameterized
                        updated += 1
                    except SQLAlchemyError as e:
                        logging.exception("Failed to update giocatori rowid=%s: %s", rowid, e)
                        s.rollback()
                else:
                    # insert new giocatori row using available columns
                    cur_cols = [
                        r[1]
                        for r in s.execute(
                            text("PRAGMA table_info(giocatori)")
                        ).fetchall()
                    ]
                    to_insert = {}
                    for k, v in [
                        ("Nome", nome),
                        ("Sq.", sqreal),
                        ("R.", (ruolo[:1].upper().replace("G", "P") if ruolo else "")),
                        ("Costo", costo),
                        ("squadra", team_name),
                        ("anni_contratto", 1),
                        ("opzione", "NO"),
                    ]:
                        if k in cur_cols:
                            to_insert[k] = v
                    if not to_insert:
                        notfound += 1
                    else:
                        cols = ", ".join([f'"{c}"' for c in to_insert.keys()])
                        qmarks = ", ".join([f":{c}" for c in to_insert.keys()])
                        try:
                            s.execute(
                                text(
                                    f"INSERT INTO giocatori ({cols}) VALUES ({qmarks})"
                                ),
                                to_insert,
                            )
                            inserted += 1
                        except SQLAlchemyError as e:
                            logging.exception("Failed to insert giocatori for %s: %s", nome, e)
                            s.rollback()

        # update team cash balances using SQL to preserve fantateam semantics
        for team_name in team_players.keys():
            r = s.execute(
                text("SELECT cassa_iniziale FROM fantateam WHERE squadra=:team"),
                {"team": team_name},
            ).fetchone()
            starting = float(r[0]) if r and r[0] is not None else 300.0
            spent_row = s.execute(
                text(
                    'SELECT COALESCE(SUM(CAST(REPLACE(REPLACE(REPLACE(COALESCE("Costo", "0"), ",", ""), "%", ""), " ", "") AS REAL)),0) as spent FROM giocatori WHERE squadra=:team AND NOT (opzione = "SI" AND anni_contratto IS NULL)'
                ),
                {"team": team_name},
            ).fetchone()
            spent = (
                float(spent_row[0]) if spent_row and spent_row[0] is not None else 0.0
            )
            new_attuale = starting - spent
            s.execute(
                text(
                    "INSERT OR REPLACE INTO fantateam(squadra, carryover, cassa_iniziale, cassa_attuale) VALUES (:t,0,:start,:att)"
                ),
                {"t": team_name, "start": starting, "att": new_attuale},
            )

        # commit ORM transaction
        s.commit()

        # record audit via ImportAudit model if requested
        if audit_info is not None:
            try:
                ia = ImportAudit(
                    filename=audit_info.get("filename"),
                    user=audit_info.get("user"),
                    inserted=inserted,
                    updated=updated,
                    aliases_created=audit_info.get("aliases_created", 0),
                    success=True,
                    message=audit_info.get("message"),
                )
                s.add(ia)
                s.commit()
            except SQLAlchemyError as e:
                logging.exception("Failed to record ImportAudit: %s", e)
                s.rollback()

    finally:
        s.close()

    return {"inserted": inserted, "updated": updated, "skipped": notfound}
